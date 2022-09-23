"""
We require 3 distinct libraries for the task. 1)openpyxl to get the alternatives of agents from an external source(xlsx in this case). 
2) collections to have Counter() function which would calculate the total scores
3) copy to make a deepcopy of agents preferences list, to be used in STV voting rule.   
"""
from collections import Counter
import collections
import openpyxl
import copy

preferences ={} #an empty dictionary which woudld be updated with agents and list of their alternatives. 
def generatePreferences(values):    
    """
    Parameters: 
    values(worksheet): values is a worksheet with data of preferences. This worksheet is used to generate preferences.
    
    Returns(dict):
    Returns a dictionary with agents as keys and a list of alternatives as their values.   
    """
    for row in range(1, values.max_row+1): 
        alternativelist =[]
        Preferenceslist =[]
        orderlist =[]
        for column in range(1, values.max_column+1):    
            alternativelist.append(values.cell(row,column).value)   
        for agent, valuations in enumerate(alternativelist, start = 1):#getting the preference order based on thier value and aligning them in a list using enumerate function.
            orderlist.append((agent,valuations))
        orderlist.sort(key =lambda x: (x[1], x[0]), reverse = True)
        for x in orderlist:
            Preferenceslist.append(x[0])
        preferences.update({row:Preferenceslist}) #Updating preference dictionary with agent and their alternatives in most to leat preferred order. 
    return preferences

#generatePreferences(vSheet)
def tieBreakFunction(alternatives, tieBreak,preferences): 
    """
    Parameters:
    alternatives: The list/dictionary of alternatives which had same score and need tieBreakFunction to decide the winner. 
    tieBreak: The type of tieBreak method to be used. 'max' - if maximum of tied values is to be selected. 'min' - if minimum
              agent i - checking order of tied values in agent i preference list
    preferences: list of agents and their preferred alternatives in order. 

    Returns:
    Returns the winner as per selection of max , min or agent i
    """

    if tieBreak == 'max':
        winner = max(alternatives) #If the selection is 'max', then maximum of the tied values would be considered as a winner. 
        return winner
    elif tieBreak =='min':
        winner = min(alternatives) #If the selections is 'min', the minimum of the tied values would be considered as a winner.
        return winner
    elif isinstance(tieBreak,int) :
        try:                       #If the selection is an agent number then the preference which comes first in his/her list among the tied
            if tieBreak in preferences: #values would be considered as a winner.
                newlist = []
                for i in alternatives:
                    newlist.append(preferences[tieBreak].index(i))
                return preferences[tieBreak][min(newlist)]
            else:
                raise Exception
        except Exception:        #Error message to be printed in case the integer doesnt belong to the preference list.
            print("Input integer doesnt correspond to any agent")

def dictatorship(preferences,agent):
    """
    
    """
    try:                         # We have error handling for dictatorship to print proper error message in case the   
        if agent in preferences: # agent selected is not in preferences.
            winner = preferences[agent][0]  
            return winner  
        else:
            raise Exception
    except Exception:
        print('Input Integer doesnt correspond to any agent. Please try again')
"""
All the voting rules , except for score vector and rangevoting, take two parameters and return the winner
All of them create a scoring list and use zip() function to assign them to alternatives of the agenets. 
Parameters: 
    preferences (dict): list of agents and their preferred alternatives in order. 
    tieBreak (str/int): The type of tieBreak method to be used. 'max' - if maximum of tied values is to be selected. 'min' - if minimum
               agent i - checking order of tied values in agent i preference list

    Returns (int):
    Returns winner after calculating total scores for alternatives. If there is a tie in scores, tieBreakFunction helps decide the winner. 
""" 
def scoringRule(preferences,scorevector,tieBreak):
    """
    Parameters:
    In addition to 2 parameters specified above we have
    scorevector(list): a list of scoes to be awared to preferences of the agents. Lenght should be equal to no.of alternatives

    The maximum value in scorevector is assigned to 1st preference of an agent and least to the last. The total score is calculated 
    using Counter funciton.
    """
    scoringrules =[]   
    scorecalculation = Counter()
    scorevector.sort(reverse = True) 
    tiedvalues =[]  
    try:                                    #Error handling. to check if length of scorevector is equal to no.of alternatives. 
        for i in preferences:    
            if len(preferences[i]) == len(scorevector):         
                scoringlist = {} 
            for a,b in zip(preferences[i],scorevector): #Using zip function to assign scores from scoreslist to preferences.
                scoringlist.update({a:b})               #Appending the alternatives and their scores in a dictionary.
            scoringrules.append(scoringlist)
            for sr in scoringrules:
                scorecalculation.update(sr)             #Calculating total scores by Counter() method.
            scoredict = dict(scorecalculation)
            winner = max(scoredict, key = scoredict.get) #Generating alternatve with maximum scores
            if len(set(scoredict.values())) != len(scoredict.values()):
                tiedvalues = [key for key, value in scoredict.items() if value == max(scoredict.values())]
            if len(tiedvalues)==0:                     #tiedvalues is a list of alternatives with maximum scores.
                return winner
            if winner in tiedvalues and len(tiedvalues) == 1:
                return winner
            else:
                return tieBreakFunction(tiedvalues,tieBreak,preferences)
    except Exception:
            print("Incorret Input")    #Error message to be printed in case the lenght of scorevector is wrong. 

def plurality(preferences,tieBreak):
    """
    plurality considers the alternatives which appear the most at the first position of preference.
    """
    Pluralitylist =[]
    tiedvalues =[]
    for i in preferences:
        Pluralitylist.append(preferences[i][0])      #Appending 1st alternative of each preference list.
        winner = max(Pluralitylist, key= Pluralitylist.count)  
    occurences = collections.Counter(Pluralitylist)  #Calculating the frequency or count of alternatives appearing first.
    if len(set(occurences.values())) != len(occurences.values()):
        tiedvalues = [key for key, value in occurences.items() if value == max(occurences.values())]
    if len(tiedvalues) ==0:                          #if there are no tied values, the function retunrs the winner. 
        return winner
    elif winner in tiedvalues and len(tiedvalues) == 1: #if there is a tie but between any alterntive other than highest 
        return winner                                   #Function returns the winner again.
    else:
        return tieBreakFunction(tiedvalues,tieBreak,preferences) #Calling tirBreakFunction if there is a tie between maximum.

def veto(preferences,tieBreak):
    """
    Veto voting rule gives 0 score to alternative at the last position. Rest all are awarded score of 1. 
    """
    scorelist = []
    vetolist =[]
    tiedvalues =[]
    scorecalculation = Counter()
    for i in preferences:              #Generating a list with scores to be assigned later to alternatives list 
        while len(scorelist)< len(preferences[i])-1:
            scorelist.append(1)
    scorelist.append(0)
    for i in preferences:
        vetodic ={}
        for a,b in zip(preferences[i],scorelist): 
            vetodic.update({a:b})      
        vetolist.append(vetodic)
    for vt in vetolist:
        scorecalculation.update(vt)   
    scoredict = dict(scorecalculation)  
    winner = max(scoredict, key = scoredict.get)
    if len(set(scoredict.values())) != len(scoredict.values()):
        tiedvalues = [key for key, value in scoredict.items() if value == max(scoredict.values())]
    if len(tiedvalues) ==0:
        return winner
    elif winner in tiedvalues and len(tiedvalues) == 1:
        return winner
    else:
        return tieBreakFunction(tiedvalues,tieBreak,preferences)

def borda(preferences,tieBreak):
    """
    Borda voting rules assigns 0 to least preferred alternative and for the rest score increases by 1 as position decreases. 
    """
    scorelist = []
    bordalist =[]
    tiedvalues =[]
    scorecalculation = Counter()
    for i in preferences:
        a = len(preferences[i])-1                 
        while len(scorelist) < len(preferences[i]):  #Generating a list of scores by using length of preferences. 
            scorelist.append(a)
            a-=1
    for i in preferences:
        bordadic ={}
        for a,b in zip(preferences[i], scorelist): #Using the generated list to score alternatives of agents
            bordadic.update({a:b})
        bordalist.append(bordadic)
    for bd in bordalist:
        scorecalculation.update(bd)
    scoresdict = dict(scorecalculation)
    winner = max(scoresdict, key = scoresdict.get)
    if len(set(scoresdict.values())) != len(scoresdict.values()):
        tiedvalues = [key for key, value in scoresdict.items() if value == max(scoresdict.values())]
    if len(tiedvalues) ==0:
        return winner 
    elif winner in tiedvalues and len(tiedvalues) == 1:
        return winner
    else:
       return tieBreakFunction(tiedvalues,tieBreak,preferences)

def harmonic(preferences,tieBreak):
    """
    Harmonic rule assigns score of 1/j where j is the position of alternative in the preference list.  
    """
    harmoniclist =[]  
    list1 =[]  
    scoredict ={}
    tiedvalues =[]
    scorecalculation =Counter()
    for i in preferences:
        a = 1/(len(preferences[i]))
        while len(harmoniclist) < len(preferences[i]): #Similar to Borda, using the no.of alternatives to generate scoring list 
            a = round(a,2)
            harmoniclist.append(a)
            if  len(preferences[i]) != len(harmoniclist):
                a = 1/(len(preferences[i])-len(harmoniclist))            
                exit
    harmoniclist.reverse()                  #Reversing the list of scores so that first preference gets the max score and last one gets minimum 
    for i in preferences:
        harmonicdic = {}
        for a,b in zip(preferences[i],harmoniclist):
            harmonicdic.update({a:b})
        list1.append(harmonicdic)
    for hm in list1:
        scorecalculation.update(hm)
    scoredict = dict(scorecalculation)
    winner = max(scoredict, key = scoredict.get)
    if len(set(scoredict.values())) != len(scoredict.values()):
        tiedvalues = [key for key, value in scoredict.items() if value == max(scoredict.values())]
    if len(tiedvalues) ==0:
        return winner 
    elif winner in tiedvalues and len(tiedvalues) == 1:
        return winner
    else:
       return tieBreakFunction(tiedvalues,tieBreak,preferences)

def STV(preferences,tieBreak):
    """
    STV works in rounds where in each round the alternative which appears least no.of times at the first position is removed. 
    """
    STVdict= copy.deepcopy(preferences)
    for i in STVdict:
        STVlist =[]
        while len(STVdict[i])>1:                 #Since STV works in rounds, using the condition of lenght of alternatives list. 
            for i in STVdict:               
                STVlist.append(STVdict[i][0])    #Appending the alternatives which appear first. 
                occurences = collections.Counter(STVlist)  
                for j in STVdict[i]:
                    if j not in STVlist: 
                        occurences.update({j:0})  #If any alternative doesn't appear first, appending them to list with 0 score.
                mini = min(occurences, key= occurences.get) 
            for i in STVdict: 
                STVdict[i].remove(mini)         #Deleting the least occuring alternatives from copied dictionary 
                STVlist.clear()                 #Clearing the list for the next round  
            occurences.pop(mini)                #Deleting the alternative from scoring dictionary for the next round
    if len(occurences)>1:
        return tieBreakFunction(occurences,tieBreak,preferences)   
    else:
        return STVdict[i][0]   

def rangeVoting(values,tieBreak):
    """
    Parameters:
    values(xlsx): values is a worksheet with data of preferences. This worksheet is used to generate preferences. 

    Rangevoting takes in values from worksheet and calculates the total sum of valuations. 
    """
    rangevotinglist =[]
    tiedvalues =[]
    for row in range(1, values.max_row+1):
        alternativelist =[]  
        tiedvalues =[]     
        scorecalculation =Counter()
        for column in range(1, values.max_column+1):   
            srlist =[]
            srdic ={} 
            alternativelist.append(values.cell(row,column).value) 
        for agent, valuations in enumerate(alternativelist, start = 1):    
            srlist.append((agent,valuations))
            srdic.update(srlist)     
        rangevotinglist.append(srdic)
    for rv in rangevotinglist:
        scorecalculation.update(rv)         #Calculating the sum of valuations from the raw the file.
    scoresdict = dict(scorecalculation)
    winner = max(scoresdict, key = scoresdict.get)
    if len(set(scoresdict.values())) != len(scoresdict.values()):
        tiedvalues = [key for key, value in scoresdict.items() if value == max(scoresdict.values())]
    if len(tiedvalues) ==0:
        return winner
    elif winner in tiedvalues and len(tiedvalues) == 1:
        return winner
    else:
       return tieBreakFunction(tiedvalues,tieBreak,preferences)

